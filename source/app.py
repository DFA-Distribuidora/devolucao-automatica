from openpyxl import *
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

#carregar o workbook e a worksheet
arquivo = 'devolucao.xlsx'
wb = load_workbook(arquivo)
ws = wb.active

array_cod_prod = []
array_nome_prod = []
array_quant_result = []
array_lote = []

#guardar o array de codigos de produtos
for row in range(2, 6):
    array_cod_prod.append(ws['A'+str(row)].value)

#guardar o array de nomes de produtos
for row in range(2, 6):
    array_nome_prod.append(ws['B'+str(row)].value)

#guardar o array de quantidades resultantes de produtos
for row in range(2, 6):
    array_quant_result.append(int(ws['C'+str(row)].value)-int(ws['F'+str(row)].value))

#guardar o array de lotes de produtos
for row in range(2, 6):
    array_lote.append(ws['D'+str(row)].value)

driver = webdriver.Firefox()
driver.get("http://10.250.1.8:8090/Core/StockMovement/ProductTransfer")
#elem = driver.find_element(By.ID, "Login")
#elem.send_keys("0028")
#elem = driver.find_element(By.ID, "Password")
#elem.send_keys("Claudiane.1")
#elem.send_keys(Keys.RETURN)
#elem.find_element(By.XPATH, "//*[text()='Processos Internos']")
#elem.click()

#element = WebDriverWait(driver, 10).until(
#        EC.presence_of_element_located((By.ID, "myDynamicElement"))
#    )


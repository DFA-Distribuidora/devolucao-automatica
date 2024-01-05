from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

#createing an workbook
wb = Workbook() #cria o wb
ws = wb.active
# ws2 = wb.create_sheet('MySheet2')
ws.title = 'MySheet' #Defina o nome da sheet
wb.save('Teste.xlsx') #salva o workbook

#Loading an existing workbook
# wb = load_workbook('exemplo.xlsx') #workbook é o mesmo que um livro de planilhas (o arquivo xlsx)

#Accessing worksheets
# ws = wb.active # uma worksheet sao as planilhas pertencente ao livro de planilhas (o arquivo xlsx)
# ws = wb['<Qualquer outra sheet sem precisar ser a sheet ativa>']

# print(ws)

#Accessing cell values
# celula = ws['B1'].value
# print(celula)

#Changing the value of a cell
# ws['C1'] = 'Notas'
# print(ws['C1'].value)

#Saving workbooks
# wb.save('exemplo.xlsx')

#Creating, listing and changing sheets
# print(wb.sheetnames) # mostra todas as sheets
# print(wb['Planilha1']) #mostra uma sheet em especifico
# wb.create_sheet('Planilha2') #criar uma sheet no workbook

#Creating a new Workbook
# arquivo = 'exemplo.xlsx'
# wb = load_workbook(arquivo)
# ws = wb.active
# ws.append(['Tim', 'Is', 'Great', '!'])#putting some informations (adding/appeding rows)
# wb.save('Minha planilha.xlsx')

#Accessing multiples cells
# for row in range(1, 11):
#     for col in range(1, 5): #manualmente seria: for col in range(0, 4) por causa do 65 ja setado que vale A
#         # char = chr(65 + col) #pegar a letra da coluna manualmente
#         char = get_column_letter(col)
#         print(ws[char + str(row)].value)

#Merging cells
# ws.merge_cells("A5:D5") #merge
# ws.unmerge_cells("A5:D5") #unmerge
# wb.save('exemplo.xlsx')

#Inserting and Deleting Rows and Columns
# ws.insert_rows(7) #insere uma linha vazia na linha 7
# ws.delete_rows(7) #exclui a linha 7
# ws.insert_cols(2) #insere uma coluna no lugar da coluna 'B'
# ws.delete_cols(2) #exclui a coluna 'B'

#Copying and Moving Cells
# rows positivas -> abaixo | rows negativas -> cima
# cols positivas -> direita | cols negativas -> esquerda
# ws.move_range("A1:C3", rows=3, cols=3)
# wb.save(arquivo)
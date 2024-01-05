from time import sleep
from openpyxl import *
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from tkinter import filedialog as fd


class Produtos:
        
    def __init__(self):
        self.tamanho_devolucao = 0
        self.linha = 2
        self.array_cod_prod = []
        self.array_nome_prod = []
        self.array_quant_result = []
        self.array_lote = []
        
    def set_planilha(self):
        arquivo = fd.askopenfilename(
            initialdir='C:\\Users\\Win\\Desktop\\devolucao_automatic\\',
            title="Escolha a planilha de devolução",
            filetypes=(('Planilha do Excel', '*.xlsx'), ("all files", '*.*'))
        )
        self.wb = load_workbook(arquivo)
        self.ws = self.wb.active
    
    def set_codigos_produtos(self):
        #guardar o array de codigos de produtos
        for row in range(2, self.tamanho_devolucao+2):
            self.array_cod_prod.append(self.ws['A'+str(row)].value)

    def set_nomes_produtos(self):
        #guardar o array de nomes de produtos
        for row in range(2, self.tamanho_devolucao+2):
            self.array_nome_prod.append(self.ws['B'+str(row)].value)
            
    def set_quantidades(self):    
        #guardar o array de quantidades resultantes de produtos
        for row in range(2, self.tamanho_devolucao+2):
            valor_c = self.ws['C'+str(row)].value
            valor_f = self.ws['F' + str(row)].value
            valor_c = 0 if valor_c is None else int(valor_c)
            valor_f = 0 if valor_f is None else int(valor_f)
            self.array_quant_result.append(valor_c - valor_f)

    def set_lotes_produtos(self):
        #guardar o array de lotes de produtos
        for row in range(2, self.tamanho_devolucao+2):
            self.array_lote.append(str(self.ws['D'+str(row)].value))

    def set_tam_dev(self):
        #saber o tamanho da devolucao
        while self.ws['A'+str(self.linha)].value != None:
            self.linha +=1
            self.tamanho_devolucao +=1
            if self.ws['A'+str(self.linha)].value == None:
                break
            
    def set_tudo(self):
        self.set_planilha()
        self.set_tam_dev()
        self.set_codigos_produtos()
        self.set_lotes_produtos()
        self.set_nomes_produtos()
        self.set_quantidades()
        
    def enviar_dados(self):
        return self